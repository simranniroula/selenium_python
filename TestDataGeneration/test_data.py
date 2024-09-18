from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

fake_data = Faker()
# print(fake_data.name())

for i in range(1, 10):
    for j in range(1, 4):
        ws.cell(row=i, column=1).value = fake_data.name()
        ws.cell(row=i, column=1).value = fake_data.email()
        ws.cell(row=i, column=1).value = fake_data.address()
wb.save("test_data.xlsx")

